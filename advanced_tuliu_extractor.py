import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import random
from datetime import datetime, timedelta
import re
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException
)
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
import json

# 配置文件路径，用于记忆保存目录
CONFIG_PATH = "config.json"


def load_config():
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_config(config):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except:
        pass


class AdvancedTuliuDetailExtractor:
    """
    土流网爬虫 —— 仅保存筛选后符合要求的数据
    """

    def __init__(self, log_callback=None):
        self.driver = None
        self.wait = None
        self.original_window = None
        self.all_raw_data = []
        self.log_callback = log_callback
        self.is_running = False
        self.land_type = "建设用地"
        self.save_dir = ""

        self.land_type_xpath = {
            "建设用地": "/html/body/div[2]/div[3]/div[1]/div[2]/dl[2]/dd/ul/li[3]/a",
            "林地": "/html/body/div[2]/div[3]/div[1]/div[2]/dl[3]/dd/ul/li[3]/a"
        }

    def set_save_dir(self, save_dir):
        self.save_dir = save_dir

    def set_land_type(self, land_type):
        if land_type in self.land_type_xpath:
            self.land_type = land_type
            self.log(f"✅ 已选择爬取 {land_type} 数据")

    def log(self, msg):
        timestamp = datetime.now().strftime("[%Y-%m-%d %H:%M:%S]")
        full_msg = f"{timestamp} {msg}"
        print(full_msg)
        if self.log_callback:
            self.log_callback(full_msg)

    def setup_driver(self):
        try:
            self.log("正在初始化浏览器驱动...")
            service = Service(ChromeDriverManager().install())
            options = webdriver.ChromeOptions()

            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-gpu')
            options.add_argument('--start-maximized')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)

            prefs = {
                "profile.managed_default_content_settings.images": 2,
                "profile.default_content_setting_values.notifications": 2
            }
            options.add_experimental_option("prefs", prefs)

            self.driver = webdriver.Chrome(service=service, options=options)
            self.wait = WebDriverWait(self.driver, 15)

            self.driver.execute_script("""
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            """)
            self.log("✅ 浏览器驱动设置成功")
        except Exception as e:
            self.log(f"❌ 驱动设置失败: {e}")
            raise

    def random_sleep(self, min_s=1, max_s=3):
        time.sleep(random.uniform(min_s, max_s))

    def visit_find_land_page(self):
        try:
            self.log("🌐 访问广东土流网...")
            self.driver.get("https://guangdong.tuliu.com/")
            self.original_window = self.driver.current_window_handle
            self.random_sleep(3, 5)

            try:
                guangzhou_btn = WebDriverWait(self.driver, 20).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "/html/body/div[3]/div[2]/div[1]/div[2]/div[1]/dl[1]/dd/a[3]"))
                )
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", guangzhou_btn)
                self.random_sleep(1, 2)
                self.driver.execute_script("arguments[0].click();", guangzhou_btn)
                self.log("✅ 精准定位并点击 '广州' 链接成功")
                guangzhou_clicked = True
            except Exception as e:
                self.log(f"❌ 精准定位广州链接失败: {e}")
                guangzhou_clicked = False

            if not guangzhou_clicked:
                self.log("❌ 未找到广州相关可点击元素")
                return False

            self.random_sleep(3, 5)
            try:
                WebDriverWait(self.driver, 10).until(lambda d: len(d.window_handles) > 1)
                for h in self.driver.window_handles:
                    if h != self.original_window:
                        self.driver.switch_to.window(h)
                        self.log("✅ 已切换到广州页面新标签页")
                        break
            except:
                self.log("✅ 广州页面在当前标签页打开，无需切换")

            try:
                land_xpath = self.land_type_xpath[self.land_type]
                land_btn = WebDriverWait(self.driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, land_xpath))
                )
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", land_btn)
                self.random_sleep(1, 2)
                self.driver.execute_script("arguments[0].click();", land_btn)
                self.log(f"✅ 精准定位并点击 '{self.land_type}' 链接成功")
                self.random_sleep(3, 5)
            except Exception as e:
                self.log(f"❌ 点击 {self.land_type} 失败: {e}")
                return False

            return True
        except Exception as e:
            self.log(f"❌ 访问页面失败: {e}")
            return False

    def collect_all_page_links(self, max_page=5):
        all_links = []
        current_page = 1

        self.log(f"\n🚀 开始收集 {max_page} 页的 {self.land_type} 链接...")

        while current_page <= max_page and self.is_running:
            self.log(f"\n📄 收集第 {current_page} 页链接")
            page_links = self._get_one_page_links()
            if page_links:
                all_links.extend(page_links)
                self.log(f"✅ 第 {current_page} 页收集到 {len(page_links)} 条链接")
            else:
                self.log(f"⚠️ 第 {current_page} 页无链接，停止收集")
                break

            if not self._click_next_page():
                self.log(f"🔚 第 {current_page} 页已是最后一页，停止收集")
                break

            current_page += 1
            self.random_sleep(3, 5)

        self.log(f"\n🏁 链接收集完成！共收集到 {len(all_links)} 条 {self.land_type} 链接")
        return all_links

    def _get_one_page_links(self):
        try:
            self.wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.col-sm-4")))
            items = self.driver.find_elements(By.CSS_SELECTOR, "div.col-sm-4")
            links = []
            for item in items:
                try:
                    a = item.find_element(By.TAG_NAME, "a")
                    url = a.get_attribute("href")
                    title = a.get_attribute("title") or a.text.strip()
                    if url and ".html" in url and title:
                        links.append({"详情链接": url, "标题": title})
                except:
                    continue
            return links
        except Exception as e:
            self.log(f"❌ 提取本页链接失败: {e}")
            return []

    def _click_next_page(self):
        try:
            next_btn = self.driver.find_element(By.CSS_SELECTOR, "a.next.btn.btn-bg")
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_btn)
            self.random_sleep(1, 2)
            self.driver.execute_script("arguments[0].click();", next_btn)
            self.wait.until(EC.staleness_of(next_btn))
            self.random_sleep(3, 5)
            return True
        except NoSuchElementException:
            try:
                next_btn = self.driver.find_element(By.XPATH,
                                                    "//a[contains(text(), '下一页') and contains(@class, 'btn')]")
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_btn)
                self.driver.execute_script("arguments[0].click();", next_btn)
                self.wait.until(EC.staleness_of(next_btn))
                self.random_sleep(3, 5)
                return True
            except:
                return False
        except:
            return False

    def crawl_all_details(self, link_list):
        self.all_raw_data = []
        total = len(link_list)

        self.log(f"\n🚀 开始爬取 {total} 条 {self.land_type} 链接的核心详情数据...")

        for i, link in enumerate(link_list):
            if not self.is_running:
                self.log("⚠️ 爬虫已被手动停止")
                break

            self.log(f"\n[{i + 1}/{total}] 处理: {link['标题'][:40]}")
            detail = self._extract_detail(link["详情链接"])
            if detail:
                core_data = {
                    "发布时间": detail["发布时间"],
                    "土地类型": detail["土地类型"],
                    "地点": detail["地点"],
                    "面积": detail["面积"],
                    "价格": detail["价格"],
                    "电话": detail["电话"],
                    "姓名": detail["姓名"]
                }
                self.all_raw_data.append(core_data)
            self.random_sleep(2, 4)

        self.log(f"\n🏁 详情爬取完成！共获取 {len(self.all_raw_data)} 条 {self.land_type} 核心数据")
        return self.all_raw_data

    def _extract_detail(self, url):
        try:
            self.driver.get(url)
            self.random_sleep(2, 4)
            data = {
                "发布时间": "",
                "土地类型": "",
                "地点": "",
                "面积": "",
                "价格": "",
                "电话": "",
                "姓名": ""
            }

            try:
                el = self.driver.find_element(By.XPATH, "//span[contains(text(),'更新时间')]")
                raw_time = el.text.replace("更新时间：", "").replace("更新时间:", "").strip()
                data["发布时间"] = raw_time
            except:
                pass

            try:
                el = self.driver.find_element(By.XPATH,
                                              "//div[contains(text(),'土地类型')]/following-sibling::div[@class='col-sm-9 text-gray-4']")
                data["土地类型"] = el.text.strip()
            except:
                pass

            try:
                location = self.driver.find_element(
                    By.XPATH, "//div[text()='土地地点']/following-sibling::div[@class='col-sm-9 text-gray-4']"
                ).text.strip()
                data["地点"] = location
            except:
                data["地点"] = ""

            try:
                el = self.driver.find_element(By.XPATH,
                                              "//div[contains(text(),'土地面积')]/following-sibling::div[@class='col-sm-2 text-gray-4']")
                data["面积"] = el.text.strip()
            except:
                pass

            try:
                price_el = self.driver.find_element(By.CSS_SELECTOR, "p.font-18.text-warning.padding-l-0")
                data["价格"] = price_el.text.strip()
                if not data["价格"]:
                    price_el = self.driver.find_element(By.XPATH,
                                                        "//div[contains(text(),'价格')]/following-sibling::div")
                    data["价格"] = price_el.text.strip()
            except:
                try:
                    price_el = self.driver.find_element(By.XPATH, "//span[@class='price']")
                    data["价格"] = price_el.text.strip()
                except:
                    data["价格"] = ""

            try:
                data["电话"] = self.driver.find_element(By.ID, "full_phone_num").get_attribute("value").strip()
            except:
                data["电话"] = ""

            try:
                contact = self.driver.find_element(By.ID, "land_contact_broker").text.strip()
                if contact:
                    data["姓名"] = contact
                else:
                    contact = self.driver.find_element(By.ID, "trader_name").text.strip()
                    data["姓名"] = contact
            except:
                try:
                    contact = self.driver.find_element(By.ID, "trader_name").text.strip()
                    data["姓名"] = contact
                except:
                    data["姓名"] = ""

            return data
        except Exception as e:
            self.log(f"❌ 详情提取失败: {url} | {e}")
            return None

    def _clean_and_parse_time(self, time_str):
        if not time_str:
            return None

        clean_str = re.sub(r'^更新时间[:：]?', '', time_str).strip()
        time_formats = [
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d", "%Y年%m月%d日 %H:%M:%S", "%Y年%m月%d日"
        ]

        for fmt in time_formats:
            try:
                return datetime.strptime(clean_str, fmt)
            except:
                continue

        self.log(f"⚠️ 无法解析的时间字符串: {clean_str} | 原始: {time_str}")
        return None

    def filter_recent_data(self, days=30):
        if not self.all_raw_data:
            self.log("⚠️ 无数据可筛选")
            return []

        self.log(f"\n🚀 开始筛选最近 {days} 天更新的 {self.land_type} 数据...")
        filtered_data = []
        cutoff_date = datetime.now() - timedelta(days=days)

        for data in self.all_raw_data:
            update_time = data.get("发布时间", "")
            if not update_time:
                continue

            update_date = self._clean_and_parse_time(update_time)
            if not update_date:
                continue

            if update_date >= cutoff_date:
                filtered_data.append(data)
                self.log(f"✅ 保留: {data['地点'][:20]} | {update_time} | 价格: {data['价格']}")
            else:
                self.log(f"❌ 过滤: {data['地点'][:20]} | {update_time} | 价格: {data['价格']}")

        self.log(
            f"\n🏁 筛选完成！符合条件的 {self.land_type} 数据共 {len(filtered_data)} 条")
        return filtered_data

    def save_to_xlsx(self, filtered_data, filter_days):
        if not filtered_data:
            self.log(f"⚠️ 无符合条件的 {self.land_type} 数据，跳过导出")
            return

        fields = [
            "发布时间", "土地类型", "地点", "面积", "价格", "电话", "姓名"
        ]

        column_widths = {
            "发布时间": 22,
            "土地类型": 20,
            "地点": 25,
            "面积": 15,
            "价格": 18,
            "电话": 20,
            "姓名": 12
        }

        wb = Workbook()
        ws = wb.active
        ws.title = "筛选数据"
        ws.append(fields)

        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_data in filtered_data:
            ws.append([row_data.get(field, "") for field in fields])

        for col_idx, field in enumerate(fields, 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = column_widths[field]

        filename = f"土流网_最近{filter_days}天{self.land_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        save_path = os.path.join(self.save_dir, filename)
        wb.save(save_path)
        self.log(f"💾 文件已保存：{save_path}")

    def close(self):
        self.is_running = False
        if self.driver:
            self.driver.quit()
            self.log("\n🔌 浏览器已关闭")

    def run_spider(self, max_page, filter_days):
        self.is_running = True
        try:
            self.setup_driver()
            if self.visit_find_land_page() and self.is_running:
                all_links = self.collect_all_page_links(max_page=max_page)
                if all_links and self.is_running:
                    raw_data = self.crawl_all_details(all_links)
                    filtered_data = self.filter_recent_data(days=filter_days)
                    self.save_to_xlsx(filtered_data, filter_days)

            self.log(f"\n🎉 {self.land_type} 任务完成！")
        except Exception as e:
            self.log(f"\n💥 出错：{e}")
        finally:
            self.close()


class TuliuSpiderGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("土流网爬虫")
        self.root.geometry("800x600")
        self.config = load_config()
        self.spider = None
        self.spider_thread = None
        self.selected_land_type = tk.StringVar(value="建设用地")

        self._create_widgets()

    def _create_widgets(self):
        config_frame = ttk.LabelFrame(self.root, text="配置")
        config_frame.pack(padx=10, pady=10, fill=tk.X)

        # 土地类型（修复 pad → padx）
        land_frame = ttk.Frame(config_frame)
        land_frame.grid(row=0, column=0, columnspan=4, padx=5, pady=5, sticky="w")
        ttk.Label(land_frame, text="类型：").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(land_frame, text="建设用地", variable=self.selected_land_type, value="建设用地").pack(
            side=tk.LEFT, padx=5)
        ttk.Radiobutton(land_frame, text="林地", variable=self.selected_land_type, value="林地").pack(side=tk.LEFT,
                                                                                                      padx=5)

        # 页数（修复 pad → padx）
        ttk.Label(config_frame, text="爬取页数：").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.page_entry = ttk.Entry(config_frame, width=10)
        self.page_entry.grid(row=1, column=1, padx=5, pady=5)
        self.page_entry.insert(0, "5")

        # 天数（修复 pad → padx）
        ttk.Label(config_frame, text="筛选天数：").grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.days_entry = ttk.Entry(config_frame, width=10)
        self.days_entry.grid(row=1, column=3, padx=5, pady=5)
        self.days_entry.insert(0, "30")

        # 保存路径（修复 pad → padx）
        path_frame = ttk.Frame(config_frame)
        path_frame.grid(row=2, column=0, columnspan=4, padx=5, pady=5, sticky="w")
        ttk.Label(path_frame, text="保存到：").pack(side=tk.LEFT, padx=5)
        self.path_entry = ttk.Entry(path_frame, width=45)
        self.path_entry.pack(side=tk.LEFT, padx=5)
        ttk.Button(path_frame, text="选择目录", command=self.choose_path).pack(side=tk.LEFT, padx=5)

        # 自动加载上次路径
        last_path = self.config.get("last_save_path", "")
        if last_path and os.path.isdir(last_path):
            self.path_entry.insert(0, last_path)

        # 按钮（修复 pad → padx）
        btn_frame = ttk.Frame(self.root)
        btn_frame.pack(padx=10, pady=5, fill=tk.X)
        self.start_btn = ttk.Button(btn_frame, text="开始", command=self.start)
        self.start_btn.pack(side=tk.LEFT, padx=5)
        self.stop_btn = ttk.Button(btn_frame, text="停止", command=self.stop, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="清空日志", command=self.clear_log).pack(side=tk.LEFT, padx=5)

        # 日志（修复 pad → padx）
        log_frame = ttk.LabelFrame(self.root, text="日志")
        log_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        self.log_text = scrolledtext.ScrolledText(log_frame, font=("Consolas", 9))
        self.log_text.pack(padx=5, pady=5, fill=tk.BOTH, expand=True)
        self.log_text.config(state=tk.DISABLED)

    def choose_path(self):
        path = filedialog.askdirectory(title="选择保存文件夹")
        if path:
            self.path_entry.delete(0, tk.END)
            self.path_entry.insert(0, path)
            self.config["last_save_path"] = path
            save_config(self.config)

    def append_log(self, msg):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def start(self):
        save_path = self.path_entry.get().strip()
        if not save_path or not os.path.isdir(save_path):
            messagebox.showwarning("提示", "请先选择有效的保存目录")
            return

        try:
            max_page = int(self.page_entry.get())
            days = int(self.days_entry.get())
        except:
            messagebox.showerror("错误", "页数和天数必须是数字")
            return

        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

        self.spider = AdvancedTuliuDetailExtractor(self.append_log)
        self.spider.set_save_dir(save_path)
        self.spider.set_land_type(self.selected_land_type.get())

        self.spider_thread = threading.Thread(target=self.spider.run_spider, args=(max_page, days), daemon=True)
        self.spider_thread.start()
        self.check_thread()

    def stop(self):
        if self.spider:
            self.spider.is_running = False
            self.append_log("⏹ 已停止")
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)

    def check_thread(self):
        if self.spider_thread and self.spider_thread.is_alive():
            self.root.after(200, self.check_thread)
        else:
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)

    def clear_log(self):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)


if __name__ == "__main__":
    root = tk.Tk()
    app = TuliuSpiderGUI(root)
    root.mainloop()